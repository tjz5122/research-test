import os
import shlex
import subprocess
import threading
import signal
from itertools import product
from openpyxl import Workbook


def run_command(env):
    print('=> run commands on GPU:{}', env['CUDA_VISIBLE_DEVICES'])

    while True:
        try:
            comm = command_list.pop(0)
        except IndexError:
            break

        proc = subprocess.Popen(comm, env=env)
        proc.wait()


wb = Workbook()
ws = wb.active
ws.title = "SSM result"
ws['A1'] = "count"
ws['B1'] = "network"
ws['C1'] = "channel"
ws['D1'] = "iteration"
ws['E1'] = "dataset"
ws['F1'] = "epoch"
ws['G1'] = "lr"
ws['H1'] = "wd"
ws['I1'] = "momentum"
ws['J1'] = "batchsize"
ws['K1'] = "truncate"
ws['L1'] = "significance"
ws['M1'] = "minstat"
ws['N1'] = "samplefreq"
ws['O1'] = "varmode"
ws['P1'] = "keymode"
ws['Q1'] = "trail"


ws['R1'] = "testaccu"
ws['S1'] = "totaltime"
ws['T1'] = "convergence"

wb.save('SSMtestdata.xlsx')

test_param_groups = {}
test_param_groups["lr_group"] = [10.0,1.0,0.1]
test_param_groups["wd_group"] = [0.0001,0.0005]
test_param_groups["trail_group"] = [1,2]
test_param_groups["batch_size_group"] = [128,256]
test_param_groups["data_group"] = ["cifar10","cifar100"]
test_param_groups["epochs_group"] = [120,240,360]
test_param_groups["net_group"] = ["resnet18","mgnet"]



# flexible
chosen_param_groups = ["wd_group","trail_group"]
value_list = [test_param_groups[group] for group in chosen_param_groups]
test_list = list(product(*value_list))
for i in range(len(test_list)):
    test_list[i] = tuple(list(test_list[i]) + [i+1])

command = 'python trainexcel.py --cuda --net=resnet18 --ch=256 --iter=2222 --data cifar10 --epochs=120 --lr=1 -b=128 -m=0.8 --wd={} --km loss_plus_smooth --vm bm --minstat=100 --sf=100 --trun=0.02 --sig=0.05 --trail={} --count={}'
# flexible


command_list = []
for file in test_list:
    command_list += [command.format(*file)]
command_list = [shlex.split(comm) for comm in command_list]



# List all the GPUs you have

ids_cuda = [0,1,2,3]
for c in ids_cuda:
    env = os.environ.copy()
    env['CUDA_VISIBLE_DEVICES'] = str(c)
    thread = threading.Thread(target = run_command, args = (env, ))
    thread.start()
    
