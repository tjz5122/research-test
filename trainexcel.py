import math
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.optim import Optimizer
from scipy import stats
import numpy as np
import torchvision
from timeit import default_timer as timer
import argparse
from openpyxl import load_workbook


class SSM_Optimizer(Optimizer):

    def __init__(self, params, lr=-1, momentum=0, weight_decay=0):
        # nu can take values outside of the interval [0,1], but no guarantee of convergence?
        if lr <= 0:
            raise ValueError("Invalid value for learning rate (>0): {}".format(lr))
        if momentum < 0 or momentum > 1:
            raise ValueError("Invalid value for momentum [0,1): {}".format(momentum))
        if weight_decay < 0:
            raise ValueError("Invalid value for weight_decay (>=0): {}".format(weight_decay))

        defaults = dict(lr=lr, momentum=momentum, weight_decay=weight_decay)
        super(SSM_Optimizer, self).__init__(params, defaults)


    def step(self, closure=None):
        loss = None
        if closure is not None:
            loss = closure()

        self.add_weight_decay()
        self.SSM_direction_and_update()
        return loss

    def add_weight_decay(self):
        
        for group in self.param_groups:
            weight_decay = group['weight_decay']
            for p in group['params']:
                if p.grad is None:
                    continue
                if weight_decay > 0:
                    p.grad.data.add_(weight_decay, p.data)

    def SSM_direction_and_update(self):

        for group in self.param_groups:
            momentum = group['momentum']
            for p in group['params']:
                if p.grad is None:
                    continue
                param_state = self.state[p]
                g_k = p.grad.data
                # get momentum buffer.
                if 'momentum_buffer' not in param_state:
                    buf = param_state['momentum_buffer'] = torch.zeros_like(p.data)
                    buf.mul_(momentum).add_(1.0 - momentum, g_k)
                else:
                    buf = param_state['momentum_buffer']
                    buf.mul_(momentum).add_(1.0 - momentum, g_k)
            
                p.data.add_(-group['lr'], buf)
                
                    
                    
class Bucket(object):
    def __init__(self, size, ratio, dtype, device, fixed_len=-1):
        self.size = size
        self.ratio = int(ratio)
        self.fixed_len = int(fixed_len)

        self.buffer = torch.zeros(size, dtype=dtype, device=device)
        self.count = 0          # number of elements kept in queue (excluding leaked)
        self.start = 0          # count = end - start
        self.end = 0
        self.total_count = 0    # total number of elements added (including leaked)
        self.statistic = 0
 
    def reset(self):
        self.buffer.zero_()    
        self.count = 0          
        self.start = 0
        self.end = 0
        self.total_count = 0

    def double_size(self):
        self.size *= 2
        self.buffer.resize_(self.size)

    def add(self, val):
        if self.end == self.size:               # when the end index reach size
            self.double_size()                      # double the size of buffer

        self.buffer[self.end] = val             # always put new value at the end
        self.end += 1                           # and increase end index by one

        if self.fixed_len > 0:
            if self.count == self.fixed_len:
                self.start += 1
            else:
                self.count += 1
        else:
            if self.total_count % self.ratio == 0:  # if leaky_count is multiple of ratio
                self.count += 1                         # increase count in queue by one
            else:                                   # otherwise leak and keep same count
                self.start += 1                         # increase start index by one

        self.total_count += 1                   # always increase total_count by one

        # reset start index to 0 and end index to count to save space
        if self.start >= self.count:
            self.buffer[0:self.count] = self.buffer[self.start:self.end]
            self.start = 0
            self.end = self.count

    # ! Need to add safeguard to allow compute only if there are enough entries
    def mean_std(self, mode='bm'):
        mean = torch.mean(self.buffer[self.start:self.end]).item() # mean of whole

        if mode == 'bm':        # batch mean variance
            b_n = int(math.floor(math.sqrt(self.count))) #batch size
            Yks = F.avg_pool1d(self.buffer[self.start:self.end].unsqueeze(0).unsqueeze(0), kernel_size=b_n, stride=b_n).view(-1) # batch mean
            diffs = Yks - mean
            std = math.sqrt(b_n /(len(Yks)-1))*torch.norm(diffs).item() # len(Yks) is number of batch
            dof = b_n - 1
        elif mode == 'olbm':    # overlapping batch mean
            b_n = int(math.floor(math.sqrt(self.count)))
            Yks = F.avg_pool1d(self.buffer[self.start:self.end].unsqueeze(0).unsqueeze(0), kernel_size=b_n, stride=1).view(-1)
            diffs = Yks - mean
            std = math.sqrt(b_n*self.count/(len(Yks)*(len(Yks)-1)))*torch.norm(diffs).item()
            dof = self.count - b_n
        else:                   # otherwise use mode == 'iid'
            std = torch.std(self.buffer[self.start:self.end]).item()
            dof = self.count - 1

        return mean, std, dof

    def stats_test(self, sigma, tolerance, mode='bm',step_test=1,truncate=0.02):
        mean, std, dof = self.mean_std(mode=mode)

        # confidence interval
        t_sigma_dof = stats.t.ppf(1-sigma/2., dof)
        self.statistic = std * t_sigma_dof / math.sqrt(self.count)
        print(self.statistic)
        if step_test != 0:
            self.statistic -= truncate
            
        return self.statistic < tolerance


class SSM(SSM_Optimizer):

    def __init__(self, params, lr=-1, momentum=0, weight_decay=0, 
                 drop_factor=10, significance=0.05, tolerance = 0.01, var_mode='bm',
                 leak_ratio=8, minN_stats=100, testfreq=100, samplefreq = 10, mode='loss_plus_smooth'):

        if lr <= 0:
            raise ValueError("Invalid value for learning rate (>0): {}".format(lr))
        if momentum < 0 or momentum > 1:
            raise ValueError("Invalid value for momentum [0,1): {}".format(momentum))
        if weight_decay < 0:
            raise ValueError("Invalid value for weight_decay (>=0): {}".format(weight_decay))
        if drop_factor < 1:
            raise ValueError("Invalid value for drop_factor (>=1): {}".format(drop_factor))
        if significance <= 0 or significance >= 1:
            raise ValueError("Invalid value for significance (0,1): {}".format(significance))
        if var_mode not in ['bm', 'olbm', 'iid']:
            raise ValueError("Invalid value for var_mode ('bm', 'olbm', or 'iid'): {}".format(var_mode))
        if leak_ratio < 1:
            raise ValueError("Invalid value for leak_ratio (int, >=1): {}".format(leak_ratio))
        # if minN_stats < 100:
        #     raise ValueError("Invalid value for minN_stats (int, >=100): {}".format(minN_stats))
        if testfreq < 1:
            raise ValueError("Invalid value for testfreq (int, >=1): {}".format(testfreq))

        super(SSM, self).__init__(params, lr=lr, momentum=momentum, weight_decay=weight_decay)
        # New Python3 way to call super()
        # super().__init__(params, lr=lr, momentum=momentum, nu=nu, weight_decay=weight_decay)

        # State initialization: leaky bucket belongs to global state.
        p = self.param_groups[0]['params'][0]
        if 'bucket' not in self.state:
            self.state['bucket'] = Bucket(1000, leak_ratio, p.dtype, p.device)

        self.state['lr'] = float(lr)
        self.state['momemtum'] = float(momentum)
        self.state['drop_factor'] = drop_factor
        self.state['significance'] = significance
        self.state['tolerance'] = tolerance
        self.state['var_mode'] = var_mode
        self.state['minN_stats'] = int(minN_stats)
        self.state['samplefreq'] = int(samplefreq)
        self.state['testfreq'] = int(testfreq)
        self.state['nSteps'] = 0
        self.state['loss'] = 0
        self.state['mode'] = mode
        self.state['step_test'] = 0

        # statistics to monitor
        self.state['smoothing'] = 0
        self.state['stats_x1d'] = 0
        self.state['stats_ld2'] = 0
        self.state['stats_val'] = 0
        self.state['stats_test'] = 0
        self.state['statistic'] = 0
        self.state['stats_stationary'] = 0
        self.state['stats_mean'] = 0
        self.state['truncate'] = 0.02


    def step(self, closure=None):
        """
        Performs a single optimization step.
        Arguments:
            closure (callable, optional): A closure that reevaluates model and returns loss.
        """
        loss = None
        if closure is not None:
            loss = closure()

        self.add_weight_decay()
        self.SSM_direction_and_update()
        self.state['nSteps'] += 1
        self.stats_adaptation()

        return loss

    def stats_adaptation(self):

        dk = self._gather_flat_buffer('step_buffer')
        xk1 = self._gather_flat_param() 
        gk = self._gather_flat_grad()
        
        
        if self.state['nSteps'] % self.state['samplefreq'] == 0:

            if self.state['mode'] == 'loss_plus_smooth':
                self.state['tolerance'] = 0.01
                self.state['smoothing'] = xk1.dot(gk).item() - (0.5 * self.state['lr']) * ((1 + self.state['momemtum'])/(1 - self.state['momemtum'])) * (dk.dot(dk).item())
                if self.state['step_test'] == 0:
                    self.state['stats_val'] =  self.state['loss'] + self.state['smoothing']
                else:
                    self.state['loss'] = np.log10(self.state['loss']) / np.log10(10/self.state['step_test']) 
                    self.state['stats_val'] = self.state['loss']  + self.state['smoothing']
                    
                    
            if self.state['mode'] == 'loss':
                self.state['tolerance'] = 0.005
                if self.state['step_test'] == 0:
                    self.state['stats_val'] =  self.state['loss'] 
                else:
                    self.state['loss'] = np.log10(self.state['loss']) / np.log10(10/self.state['step_test']) 
                    self.state['stats_val'] = self.state['loss']  
                
                
            if self.state['mode'] == 'sasa_plus':
                self.state['stats_x1d'] = xk1.dot(dk).item()
                self.state['stats_ld2'] = (0.5 * self.state['lr']) * (dk.dot(dk).item())
                self.state['stats_val'] = self.state['stats_x1d'] + self.state['stats_ld2']
        
            # add statistic to leaky bucket
            self.state['bucket'].add(self.state['stats_val'])
        
        # check statistics and adjust learning rate
        self.state['stats_test'] = 0
        self.state['stats_stationary'] = 0
        self.state['statistic'] = 0
        if self.state['bucket'].count > self.state['minN_stats'] and self.state['nSteps'] % self.state['testfreq'] == 0:
            stationary= self.state['bucket'].stats_test(self.state['significance'], self.state['tolerance'], self.state['var_mode'],self.state['step_test'],self.state['truncate'])
            self.state['stats_test'] = 1
            self.state['statistic'] = self.state['bucket'].statistic
            self.state['stats_stationary'] = int(stationary)
    
            # perform statistical test for stationarity
            if self.state['stats_stationary'] == 1:
                self.state['lr'] /= self.state['drop_factor']
                self.state['step_test'] += 1
                for group in self.param_groups:
                    group['lr'] = self.state['lr']
                self._zero_buffers('momentum_buffer')
                self.state['bucket'].reset()


    def _gather_flat_grad(self):
        views = []
        for group in self.param_groups:
            for p in group['params']:
                if p.grad is None:
                    view = p.data.new(p.data.numel()).zero_()
                elif p.grad.data.is_sparse:
                    view = p.grad.data.to_dense().view(-1) #take derivative of p
                else:
                    view = p.grad.data.view(-1) 
                views.append(view)
        return torch.cat(views, 0)
    
    # methods for gather flat parameters
    def _gather_flat_param(self):
        views = []
        for group in self.param_groups:
            for p in group['params']:
                view = p.data.view(-1)
                views.append(view)
        return torch.cat(views, 0)

    # method for gathering/initializing flat buffers that are the same shape as the parameters
    def _gather_flat_buffer(self, buf_name):
        views = []
        for group in self.param_groups:
            for p in group['params']:
                state = self.state[p]
                if buf_name not in state:  # init buffer
                    view = p.data.new(p.data.numel()).zero_()
                else:
                    view = state[buf_name].data.view(-1)
                views.append(view)
        return torch.cat(views, 0) 
    
    def _zero_buffers(self, buf_name):
        for group in self.param_groups:
            for p in group['params']:
                state = self.state[p]
                if buf_name in state:
                    state[buf_name].zero_()
        return None
    

    



### Design the Mgnet Network
use_cuda = torch.cuda.is_available()
print('Use GPU?', use_cuda)

##MgNet
class MgIte(nn.Module): 
    def __init__(self, A, S):
        super().__init__()
        
        self.A = A
        self.S = S

        self.bn1 =nn.BatchNorm2d(A.weight.size(0)) 
        self.bn2 =nn.BatchNorm2d(S.weight.size(0)) 
    
    def forward(self, out):
        u, f = out 
        u = u + F.relu(self.bn2(self.S(F.relu(self.bn1((f-self.A(u))))))) 
        out = (u, f)
        return out



class MgRestriction(nn.Module):
    def __init__(self, A_old, A, Pi, R):
        super().__init__()

        self.A_old = A_old
        self.A = A
        self.Pi = Pi
        self.R = R

        self.bn1 = nn.BatchNorm2d(Pi.weight.size(0))   
        self.bn2 = nn.BatchNorm2d(R.weight.size(0))    

    def forward(self, out):
        u_old, f_old = out 
        u = F.relu(self.bn1(self.Pi(u_old)))                              
        f = F.relu(self.bn2(self.R(f_old-self.A_old(u_old)))) + self.A(u)        
        out = (u,f)
        return out


class MgNet(nn.Module):
    def __init__(self, num_channel_input, num_iteration, num_channel_u, num_channel_f, num_classes):
        super().__init__()
        self.num_iteration = num_iteration
        self.num_channel_u = num_channel_u
        self.conv1 = nn.Conv2d(num_channel_input, num_channel_f, kernel_size=3, stride=1, padding=1, bias=False)
        self.bn1 = nn.BatchNorm2d(num_channel_f)        

        
        A = nn.Conv2d(num_channel_u, num_channel_f, kernel_size=3, stride=1, padding=1, bias=False)
        S = nn.Conv2d(num_channel_f, num_channel_u, kernel_size=3,stride=1, padding=1, bias=False)
        layers = []
        for l, num_iteration_l in enumerate(num_iteration):
            for i in range(num_iteration_l):
                layers.append(MgIte(A, S)) 
            setattr(self, 'layer'+str(l), nn.Sequential(*layers))

            if l < len(num_iteration)-1:
                A_old = A 
                
                A = nn.Conv2d(num_channel_u, num_channel_f, kernel_size=3,stride=1, padding=1, bias=False)
                S = nn.Conv2d(num_channel_f, num_channel_u, kernel_size=3,stride=1, padding=1, bias=False)

                Pi = nn.Conv2d(num_channel_u, num_channel_u, kernel_size=3,stride=2, padding=1, bias=False)
                R = nn.Conv2d(num_channel_f, num_channel_f, kernel_size=3, stride=2, padding=1, bias=False)
                
                
                layers= [MgRestriction(A_old, A, Pi, R)] 
        
        self.pooling = nn.AdaptiveAvgPool2d(1) 
        self.fc = nn.Linear(num_channel_u ,num_classes) 

    def forward(self, u, f):
        f = F.relu(self.bn1(self.conv1(f)))                
        if use_cuda:                                        
            u = torch.zeros(f.size(0),self.num_channel_u,f.size(2),f.size(3), device=torch.device('cuda')) 
        else:
            u = torch.zeros(f.size(0),self.num_channel_u,f.size(2),f.size(3))        
       
        
        out = (u, f) 

        for l in range(len(self.num_iteration)):
            out = getattr(self, 'layer'+str(l))(out)

        
        u, f = out       
        u = self.pooling(u) #do avg pooling
        u = u.view(u.shape[0], -1)  #reshape u batch_Size to vector
        u = self.fc(u)
        return u
    
    
    
##ResNet18
class BasicBlock(nn.Module):
    def __init__(self, in_planes, planes, stride):
        super().__init__()
        self.conv1 = nn.Conv2d(in_planes, planes, kernel_size=3, stride=stride, padding=1, bias=False)
        self.bn1 = nn.BatchNorm2d(planes)
        self.conv2 = nn.Conv2d(planes, planes, kernel_size=3, stride=1, padding=1, bias=False)
        self.bn2 = nn.BatchNorm2d(planes)

        self.shortcut = nn.Sequential()
        if stride != 1:
            self.shortcut = nn.Sequential(nn.Conv2d(in_planes, planes, kernel_size=1, stride=stride, bias=False),
                                          nn.BatchNorm2d(planes))

    def forward(self, x):
        out = F.relu(self.bn1(self.conv1(x)))
        out = self.bn2(self.conv2(out))
        out += self.shortcut(x)
        out = F.relu(out)
        return out


class ResNet(nn.Module):
    def __init__(self, block, num_blocks, num_classes=10):
        super().__init__()
        self.in_planes = 64

        self.conv1 = nn.Conv2d(3, 64, kernel_size=3, stride=1, padding=1, bias=False)
        self.bn1 = nn.BatchNorm2d(64)
        self.layer1 = self._make_layer(block, 64, num_blocks[0], stride=1)
        self.layer2 = self._make_layer(block, 128, num_blocks[1], stride=2)
        self.layer3 = self._make_layer(block, 256, num_blocks[2], stride=2)
        self.layer4 = self._make_layer(block, 512, num_blocks[3], stride=2)
        self.linear = nn.Linear(512, num_classes)

    def _make_layer(self, block, planes, num_blocks, stride):
        strides = [stride] + [1]*(num_blocks-1)
        layers = []
        for stride in strides:
            layers.append(block(self.in_planes, planes, stride))
            self.in_planes = planes
        return nn.Sequential(*layers)

    def forward(self, x):
        out = F.relu(self.bn1(self.conv1(x)))
        out = self.layer1(out)
        out = self.layer2(out)
        out = self.layer3(out)
        out = self.layer4(out)
        out = F.avg_pool2d(out, 4)
        out = out.view(out.size(0), -1)
        out = self.linear(out)
        return out  
    
    



###args
def get_args():
    parser = argparse.ArgumentParser(description='A general training enviroment for different learning methods.')

    parser.add_argument('--cuda',  action='store_true', help='use cuda')

    # For output 
    parser.add_argument('--name', default='result', type=str, help='the path to save the result')

    # For trail control
    parser.add_argument('--trail', default='0', type=str, help='trail of the same params.')

    # For methods    
    parser.add_argument('--epochs', type=int, help='epoch number', default=120)

    parser.add_argument('-b', '--batch-size', default=128, type=int, metavar='N', help='mini_batch size (default: 128)')
    
    parser.add_argument('--lr', '--learning-rate', default=1.0, type=float, metavar='LR', help='initial learning rate')
    
    parser.add_argument('-m', '--momentum', default=0.8, type=float, metavar='M', help='momentum')
    
    parser.add_argument('--weight-decay', '--wd', default=5e-4, type=float, metavar='W', help='weight decay (e.g. 5e-4)')
            
    parser.add_argument('--trun', metavar='truncate', default=0.02, type=float,  help='truncate value (default: 0.02)')
    
    parser.add_argument('--sig', metavar='significance', default=0.05, type=float,  help='significance level (default: 0.05)')
    
    parser.add_argument('--minstat', metavar='minstat', default=100, type=int, help='mini-stat (default: 100)')
    
    parser.add_argument('--samplefreq', '--sf', metavar='samplefreq', default=10, type=int, help='sampling frequency (default: 10)')

    parser.add_argument('--varmode', '--vm', metavar='variance_mode', default="bm", type=str, help='variance mode (default: bm)')
    
    parser.add_argument('--keymode', '--km', metavar='key_mode', default="loss_plus_smooth", type=str, help='key mode (default: loss_plus_smooth')
    

    # For Data    
    parser.add_argument('--data', type=str, help='cifar10, cifar100 or mnist', default='cifar10')
    
    # For models
    parser.add_argument('--net', metavar='neural_network', default='mgnet')
    
    parser.add_argument('--ch', metavar='channel_number', default=256, type=int, help='channel number (default: 256)')
    
    parser.add_argument('--iter', metavar='iter', default='2222')
    
    return parser.parse_args()


def main():
    args = get_args()  # get the arguments

    if args.net == "mgnet":
        args.name = 'net={},ch={},iter={},ds={},ep={},lr={},wd={},m={},bs={},trun={},sig={},minstat={},samplefreq={},var={},key={},t={}'
        args.name = args.name.format(args.net,
                                     args.ch,
                                     args.iter,
                                     args.data,
                                     args.epochs,
                                     args.lr,
                                     args.weight_decay,
                                     args.momentum,
                                     args.batch_size,
                                     args.trun,
                                     args.sig,
                                     args.minstat,
                                     args.samplefreq,
                                     args.varmode,
                                     args.keymode,
                                     args.trail
                                     )
    if args.net == "resnet18":
        args.name = 'net={},iter={},ds={},ep={},lr={},wd={},m={},bs={},trun={},sig={},minstat={},samplefreq={},var={},key={},t={}'
        args.name = args.name.format(args.net,
                                     args.iter,
                                     args.data,
                                     args.epochs,
                                     args.lr,
                                     args.weight_decay,
                                     args.momentum,
                                     args.batch_size,
                                     args.trun,
                                     args.sig,
                                     args.minstat,
                                     args.samplefreq,
                                     args.varmode,
                                     args.keymode,
                                     args.trail
                                     )
    #implementation
    minibatch_size = args.batch_size
    num_epochs =  args.epochs
    degree = args.ch
    if args.data == 'cifar10' or 'cifar100':
        num_channel_input = 3 # since cifar10
    num_channel_u = degree # usaually take channel u and f same, suggested value = 64,128,256,512.....
    num_channel_f = degree
    if args.data == 'cifar10':
        num_classes = 10
    if args.data == 'cifar100':
        num_classes = 100
    num_iteration = [int(i) for i in list(args.iter)] # for each layer do 1 iteration or you can change to [2,2,2,2] or [2,1,1,1]
    
    # Step 1: Define a model
    if args.net == "mgnet":
        my_model = MgNet(num_channel_input, num_iteration, num_channel_u, num_channel_f, num_classes)
    if args.net == "resnet18":
        my_model = ResNet(BasicBlock, num_iteration, num_classes=num_classes)
        
    if use_cuda:
        my_model = my_model.cuda()
    
    # Step 2: Define a loss function and training algorithm
    criterion = nn.CrossEntropyLoss()
    
    
    # Step 3: load dataset
    if args.data == 'cifar10':
        normalize = torchvision.transforms.Normalize(mean=(0.4914, 0.4822, 0.4465), std=(0.2023, 0.1994, 0.2010))
    
        transform_train = torchvision.transforms.Compose([torchvision.transforms.RandomCrop(32, padding=4),
                                                          torchvision.transforms.RandomHorizontalFlip(),
                                                          torchvision.transforms.ToTensor(),
                                                          normalize])
        transform_test  = torchvision.transforms.Compose([torchvision.transforms.ToTensor(),normalize])
        trainset = torchvision.datasets.CIFAR10(root='./data', train=True, download=True, transform=transform_train)
        trainloader = torch.utils.data.DataLoader(trainset, batch_size=minibatch_size, shuffle=True)
        testset = torchvision.datasets.CIFAR10(root='./data', train=False, download=True, transform=transform_test)
        testloader = torch.utils.data.DataLoader(testset, batch_size=minibatch_size, shuffle=False)
    
    if args.data == 'cifar100':
        normalize = torchvision.transforms.Normalize(mean=(0.5071, 0.4867, 0.4408), std=(0.2675, 0.2565, 0.2761))
    
        transform_train = torchvision.transforms.Compose([torchvision.transforms.RandomCrop(32, padding=4),
                                                          torchvision.transforms.RandomHorizontalFlip(),
                                                          torchvision.transforms.ToTensor(),
                                                          normalize])
        transform_test  = torchvision.transforms.Compose([torchvision.transforms.ToTensor(),normalize])
        trainset = torchvision.datasets.CIFAR100(root='./data', train=True, download=True, transform=transform_train)
        trainloader = torch.utils.data.DataLoader(trainset, batch_size=minibatch_size, shuffle=True)
        testset = torchvision.datasets.CIFAR100(root='./data', train=False, download=True, transform=transform_test)
        testloader = torch.utils.data.DataLoader(testset, batch_size=minibatch_size, shuffle=False)
    
    
    optimizer = SSM(my_model.parameters(), lr=args.lr, weight_decay=args.weight_decay, momentum=args.momentum, testfreq=len(trainloader), var_mode=args.varmode, minN_stats=args.minstat, mode=args.keymode, samplefreq=args.samplefreq, significance=args.sig)
    
    train_accuracy_list = []
    test_accuracy_list = []
    lr_list = []
    statistic_list = []
    avg_loss_list = []
    key_list = []
    epoch_time_list = []
    
    start = timer()
    #Step 4: Train the NNs
    # One epoch is when an entire dataset is passed through the neural network only once.
    for epoch in range(num_epochs):
        epoch_start_time = timer()
        running_loss = 0
        my_model.train()
        for i, (images, labels) in enumerate(trainloader):
            if use_cuda:
              images = images.cuda()
              labels = labels.cuda()
    
            # Forward pass to get the loss
            if args.net == "mgnet":
                outputs = my_model(0,images)   # We need additional 0 input for u in MgNet
            if args.net == "resnet18":
                outputs = my_model(images) 
            loss = criterion(outputs, labels)
            optimizer.state['loss'] = loss.item()
            # Backward and compute the gradient
            optimizer.zero_grad()
            loss.backward()  #backpropragation
            running_loss += loss.item()
            optimizer.step() #update the weights/parameters
        avg_loss_list.append(running_loss)
        
        
      # Training accuracy
        my_model.eval()
        correct = 0
        total = 0
        for i, (images, labels) in enumerate(trainloader):
            with torch.no_grad():
              if use_cuda:
                  images = images.cuda()
                  labels = labels.cuda()  
              if args.net == "mgnet":
                  outputs = my_model(0,images)   # We need additional 0 input for u in MgNet
              if args.net == "resnet18":
                  outputs = my_model(images) 
              p_max, predicted = torch.max(outputs, 1) 
              total += labels.size(0)
              correct += (predicted == labels).sum()
        training_accuracy = float(correct)/total
        train_accuracy_list.append(training_accuracy)     
        
        # Test accuracy
        correct = 0
        total = 0
        for i, (images, labels) in enumerate(testloader):
            with torch.no_grad():
              if use_cuda:
                  images = images.cuda()
                  labels = labels.cuda()
              if args.net == "mgnet":
                  outputs = my_model(0,images)   # We need additional 0 input for u in MgNet
              if args.net == "resnet18":
                  outputs = my_model(images) 
              p_max, predicted = torch.max(outputs, 1) 
              total += labels.size(0)
              correct += (predicted == labels).sum()
        test_accuracy = float(correct)/total
        test_accuracy_list.append(test_accuracy)
        statistic_list.append(optimizer.state['statistic'])
        key_list.append(tuple([optimizer.state['stats_val'],optimizer.state['loss'],optimizer.state['smoothing']]))
        current_lr = optimizer.state['lr']
        lr_list.append(current_lr)
        epoch_end_time = timer()
        epoch_time_list.append(epoch_end_time - epoch_start_time)
    end = timer()
    total_time = end - start
    
    print("complete")
    
    convergence = 0
    arglist = [args.net,args.ch,args.iter,args.data,args.epochs,args.lr,args.weight_decay,args.momentum,args.batch_size,args.trun,args.sig,args.minstat,args.samplefreq,args.varmode,args.keymode,test_accuracy_list[-1],total_time,convergence]
    wb = load_workbook('SSMtestdata.xlsx')
    ws = wb.active

    for col in range(len(arglist)):
        ws.cell(row=args.count+1, column=col+1, value=arglist[col+1])
    wb.save('SSMtestdata.xlsx')
        
main()

